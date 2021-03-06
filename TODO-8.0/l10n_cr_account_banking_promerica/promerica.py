# -*- encoding: utf-8 -*-
##############################################################################
#
#    Copyright (C) 2011 credativ Ltd (<http://www.credativ.co.uk>).
#    All Rights Reserved
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################<

import re
import base64
import logging
import StringIO
from datetime import datetime
from openpyxl import load_workbook
from account_banking_ccorp.parsers import models
from promerica_parser import PromericaParser
from openerp.osv import osv, fields
from openerp.tools.translate import _

bt = models.mem_bank_transaction
logger = logging.getLogger( 'l10n_cr_account_banking_promerica' )

class transaction(models.mem_bank_transaction):

    mapping = {
        'execution_date' : '',
        'effective_date' : '',
        'local_currency' : '',
        'transfer_type' : '',
        'reference' : '',
        'message' : '',
        'name' : '',
        'amount': '',
        'creditmarker': '',
    }

    def __init__(self, record, *args, **kwargs):
        '''Transaction creation'''
        #record is a dictionary, that is the reason to use iteritems().
        super(transaction, self).__init__(*args, **kwargs)
        for key, value in record.iteritems():
            if record.has_key(key):
                setattr(self, key, record[key])

        if not self.is_valid():
            logger.info("Invalid: %s", record)
    
    def is_valid(self):
        '''We don't have remote_account so override base'''
        return (self.execution_date and True) or False

class statement(models.mem_bank_statement):
    '''Bank statement imported data    '''
      
    def _transmission_number(self, record):
        self.id = record['transref']
    
    def _account_number(self, record):
        self.local_account = record['account_number']
        self.local_currency = record['currencycode']

    def _statement_number(self, record):
        self.id = record['id']
        
    def _opening_balance(self, record):
        self.start_balance = float(record['startingbalance'])
    
    def _closing_balance(self, record):
        self.end_balance = float(record['endingbalance'])
        self.date = record['bookingdate']
     
    def _transaction_new(self, record):
        parser = PromericaParser()
        sub_record = parser.statement_lines(record)
        for sub in sub_record:
            self.transactions.append(transaction(sub))

    def _not_used():
        logger.info("Didn't use record: %s", record)
        
    def _forward_available(self, record):
        self.end_balance =  float(record['endingbalance'])
        self.date = record['bookingdate'] 
    
    def _execution_date_transferred_amount (self, record):        
        self.execution_date = record['bookingdate']       
        self.transferred_amount = float(record['amount'])

    def transaction_info(self, record):
        '''Add extra information to transaction'''
        # Additional information for previous transaction
        if len(self.transactions) < 1:
            logger.info("Received additional information for non existent transaction:")
            logger.info(record)
        else:
            transaction = self.transactions[-1]
            transaction.id = ','.join([record[k] for k in ['infoline{0}'.format(i) for i in range(2,5)] if record.has_key(k)])

def raise_error(message, line):
    raise osv.osv.except_osv(_('Import error'),
        _('Error in import:%s\n\n%s') % (message, line))

class parser_promerica(models.parser):
    
    '''
        This adds a new parser in the selection options. 
        When the account is associated to a parser, the following code makes it appear as an option
    '''
    code = 'Promerica-Parser'
    name = _('Promerica Bank Statement Importer')
    country_code = 'CR'
    doc = _('''Importer for Promerica's xslx bank statements''')

    def parse(self, cr, statements_file, **kwargs):
        result = []
        parser = PromericaParser() 
        stmnt = statement()
        
        # Decode the base64 encoded file
        data = base64.decodestring(statements_file)
        
        # Gather IO from the decoded string
        file = StringIO.StringIO()
        file.write(data)
        
        # Load workbook from IO
        workbook = load_workbook(file)
        worksheet = workbook.get_active_sheet() # Get the first sheet
        
        # General Information
        account_number = worksheet.cell('B7').value # TODO: setup configuration for account number cell
        starting_balance = worksheet.cell('G13').value
        date_fields = worksheet.cell('A5').value
        
        try:
            # TODO: gather format from configuration
            # Get dates using Regular Expression
            expr = re.compile('[0-9][0-9]\/[0-9][0-9]\/[0-9][0-9][0-9][0-9]')
            dates = expr.findall(date_fields)
            
            # Format date '%Y-%m-%d'
            date_from_str = datetime.strftime(datetime.strptime(dates[0],'%d/%m/%Y'),'%Y-%m-%d')
            date_to_str = datetime.strftime(datetime.strptime(dates[1],'%d/%m/%Y'),'%Y-%m-%d')
        except:
            logger.error('Not able to read dates from file, this will possibly fail.')
            date_from_str = False
            date_to_str = False
        
        data_rows = worksheet.rows[13:]
        data = []
        for row in data_rows:
            if reduce(lambda result, value: result and (value is None or value == ''),
                      [cell.value for cell in row], True):
                break
            data.append([cell.value for cell in row])
        
        ending_balance = starting_balance
        if data:
            ending_balance = data[len(data)-1][6]
        
        # Update kwargs with read values
        kwargs.update({
                       'account_number': account_number,
                       'starting_balance': starting_balance,
                       'ending_balance': ending_balance,
                       'date_from_str': date_from_str,
                       'date_to_str': date_to_str,
                       })
        # Parse the data for the header of the stament.amount
        records = parser.parse_stamenent_record(data, **kwargs)        
        
        stmnt._transmission_number(records)
        stmnt._account_number(records)
        stmnt._statement_number(records)
        stmnt._opening_balance(records)
        stmnt._closing_balance(records)
        stmnt._forward_available(records)
        stmnt._execution_date_transferred_amount (records)
        stmnt._transaction_new(data)
        
        #call the method statement_lines in
        #parser to parse all the lines in file and add to stament.
                  
        '''
        A stament must have a header and transacctions. The method
        parse_stamenent_record parse the header and the 
        method _transaction_new parse all the line 
        (transactions) in the file. 
        '''
        if stmnt.is_valid():
            result.append(stmnt)
                      
        return result
      
# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
